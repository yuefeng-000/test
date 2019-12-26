#提示：
#1.分析数据存在哪里（打开“检查”工具，刷新页面，查看第0个请求，看【response】）
#2.观察网址规律（多翻几页，看看网址会有什么变化）
#3.获取、解析和提取数据（需涉及知识点：queue、gevent、request、BeautifulSoup、find和find_all）
#4.存储数据（csv本身的编码格式是utf-8，可以往open（）里传入参数encoding='utf-8'。这样能避免由编码问题引起的报错。)
#注：在练习的【文件】中，你能找到自己创建的csv文件。将其下载到本地电脑后，请用记事本打开，因为用Excel打开可能会因编码问题出现乱码。
from gevent import monkey
monkey.patch_all()
import gevent,requests,time
from gevent.queue import Queue
from bs4 import BeautifulSoup
import openpyxl
from openpyxl import load_workbook

open_wb=openpyxl.Workbook()
sheet=open_wb['Sheet']

wb=openpyxl.Workbook()
sheet1=wb.active
sheet1.title='sheet1'

sheet1['A1']='片名'
sheet1['B1']='导演'
sheet1['C1']='主演'
sheet1['D1']='介绍'



start=time.time()


url_list=['http://www.mtime.com/top/tv/top100/',
         'http://www.mtime.com/top/tv/top100/index-2.html',
         'http://www.mtime.com/top/tv/top100/index-3.html',
         'http://www.mtime.com/top/tv/top100/index-4.html',
         'http://www.mtime.com/top/tv/top100/index-5.html',
         'http://www.mtime.com/top/tv/top100/index-6.html',
         'http://www.mtime.com/top/tv/top100/index-7.html',
         'http://www.mtime.com/top/tv/top100/index-8.html',
         'http://www.mtime.com/top/tv/top100/index-9.html',
         'http://www.mtime.com/top/tv/top100/index-10.html']

work=Queue()

for url in url_list:
    work.put_nowait(url)
    
def crawler():
    while not work.empty():
        url=work.get_nowait()
        r=requests.get(url)
        r.encoding='utf-8'
        html=r.text
        soup=BeautifulSoup(html,'html.parser')
        #----------片名，导演，主演
        xinxis=soup.find_all(class_='mov_con')
        
        for i in xinxis:
            a=0
            xinxi=i.find_all(target='_blank')
            pianming=xinxi[0].text

            try:
                daoyan=xinxi[1].text
                
            except:
                a=1
                pass
            print('片名：'+pianming)
            print('导演：'+daoyan)
            zhuyans=[]
            for i1 in xinxi[2:]:
                zhuyan=i1.text
                zhuyans.append(zhuyan)
                yanyuans=",".join(zhuyans)
            print(yanyuans)
            try:#----------介绍
                jieshao=i.find(class_='mt3')
                print(jieshao.text)
            except:
                pass

            try:
                if a==1:
                    sheet1.append([pianming,' ',yanyuans,jieshao.text])
                sheet1.append([pianming,daoyan,yanyuans,jieshao.text])
            except:
               sheet1.append([pianming,daoyan,yanyuans])
       
            

tasks_list=[]
for x in range(2):
    task=gevent.spawn(crawler)
    tasks_list.append(task)
gevent.joinall(tasks_list)
end=time.time()
sheet1.column_dimensions["D"].auto_size = True
wb.save('时光电影TOP100电视剧.xlsx')
print(end-start)
